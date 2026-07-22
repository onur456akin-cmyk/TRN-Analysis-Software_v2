"""Export Manager Module"""

import numpy as np
from pathlib import Path
from typing import Dict, Optional
import csv
import logging

logger = logging.getLogger(__name__)


class ExportManager:
    """Export analysis results to various formats"""
    
    @staticmethod
    def export_csv(file_path: Path, data: Dict, metadata: Dict = None) -> None:
        """
        Export data to CSV file
        
        Args:
            file_path: Output file path
            data: Data dictionary with arrays
            metadata: Optional metadata dictionary
        """
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write metadata if provided
                if metadata:
                    for key, value in metadata.items():
                        writer.writerow([key, value])
                    writer.writerow([])  # Empty row
                
                # Write data
                for key, value in data.items():
                    if isinstance(value, np.ndarray):
                        writer.writerow([key])
                        if value.ndim == 1:
                            for v in value:
                                writer.writerow([v])
                        elif value.ndim == 2:
                            for row in value:
                                writer.writerow(row)
                    else:
                        writer.writerow([key, value])
                    writer.writerow([])  # Empty row
            
            logger.info(f"Data exported to CSV: {file_path}")
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise
    
    @staticmethod
    def export_txt(file_path: Path, data: Dict, metadata: Dict = None) -> None:
        """
        Export data to text file
        
        Args:
            file_path: Output file path
            data: Data dictionary
            metadata: Optional metadata dictionary
        """
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                # Write metadata
                if metadata:
                    f.write("=" * 80 + "\n")
                    f.write("METADATA\n")
                    f.write("=" * 80 + "\n")
                    for key, value in metadata.items():
                        f.write(f"{key}: {value}\n")
                    f.write("\n")
                
                # Write data
                f.write("=" * 80 + "\n")
                f.write("DATA\n")
                f.write("=" * 80 + "\n")
                
                for key, value in data.items():
                    f.write(f"\n{key}:\n")
                    if isinstance(value, np.ndarray):
                        if value.ndim == 1:
                            for v in value:
                                f.write(f"{v}\n")
                        elif value.ndim == 2:
                            for row in value:
                                f.write(" ".join(str(v) for v in row) + "\n")
                    else:
                        f.write(f"{value}\n")
            
            logger.info(f"Data exported to TXT: {file_path}")
        except Exception as e:
            logger.error(f"Error exporting to TXT: {e}")
            raise
    
    @staticmethod
    def export_excel(file_path: Path, data: Dict, metadata: Dict = None) -> None:
        """
        Export data to Excel file
        
        Args:
            file_path: Output file path
            data: Data dictionary
            metadata: Optional metadata dictionary
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            
            row = 1
            
            # Write metadata
            if metadata:
                ws['A1'] = "Metadata"
                ws['A1'].font = Font(bold=True, size=12)
                row = 2
                
                for key, value in metadata.items():
                    ws[f'A{row}'] = key
                    ws[f'B{row}'] = value
                    row += 1
                
                row += 1
            
            # Write data
            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                if isinstance(value, np.ndarray):
                    if value.ndim == 1:
                        for v in value:
                            ws[f'A{row}'] = float(v) if isinstance(v, (int, float, np.number)) else v
                            row += 1
                    elif value.ndim == 2:
                        for row_idx, row_data in enumerate(value):
                            for col_idx, v in enumerate(row_data):
                                ws.cell(row=row, column=col_idx+1, 
                                       value=float(v) if isinstance(v, (int, float, np.number)) else v)
                            row += 1
                else:
                    ws[f'A{row}'] = value
                    row += 1
                
                row += 1
            
            wb.save(file_path)
            logger.info(f"Data exported to Excel: {file_path}")
        except ImportError:
            logger.error("openpyxl not installed. Cannot export to Excel.")
            raise
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    @staticmethod
    def export_spectrum(file_path: Path, frequencies: np.ndarray, 
                       magnitude: np.ndarray, format: str = "csv") -> None:
        """
        Export spectrum data
        
        Args:
            file_path: Output file path
            frequencies: Frequency axis
            magnitude: Magnitude spectrum
            format: Export format (csv, txt, excel)
        """
        data = {
            'Frequency (Hz)': frequencies,
            'Magnitude (dB)': magnitude
        }
        
        if format == "csv":
            ExportManager.export_csv(file_path, data)
        elif format == "txt":
            ExportManager.export_txt(file_path, data)
        elif format == "excel":
            ExportManager.export_excel(file_path, data)
        else:
            raise ValueError(f"Unsupported format: {format}")


if __name__ == "__main__":
    print("Export Manager module loaded")
